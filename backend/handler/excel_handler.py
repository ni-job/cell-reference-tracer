from collections.abc import Iterator
from itertools import chain
from tempfile import NamedTemporaryFile
from typing import Any

import openpyxl as opxl
from openpyxl.cell.rich_text import CellRichText
from openpyxl.formula import Tokenizer
from openpyxl.formula.tokenizer import Token
from openpyxl.utils.cell import get_column_letter, coordinate_to_tuple, cols_from_range
from openpyxl.worksheet.formula import ArrayFormula
from streamlit.runtime.uploaded_file_manager import UploadedFile


class ExcelHandler:
    """
    Excelを扱うハンドラー
    """

    def __init__(self, file_bytes: bytes, file_name: str) -> None:
        self.__wb: opxl.Workbook

        with NamedTemporaryFile(suffix=file_name) as f:
            f.write(file_bytes)
            self.__wb = opxl.load_workbook(f.name)

    def sheet_names(self) -> list[str]:
        """
        Excelファイルのシート名一覧を取得する
        """

        return self.__wb.sheetnames

    def get(self, sheet_name: str, row: int, clm: int) -> (Any | str | CellRichText | None):
        """
        指定したシート、行、列のセルの値を取得する
        """

        return self.__wb[sheet_name].cell(row, clm).value

    def get_column_letter(self, clm: int) -> str:
        """
        列番号をアルファベットに変換する
        """
        return get_column_letter(clm)

    def coordinate_to_tuple(self, cell: str) -> tuple[int, int]:
        """
        セル座標の"A1"表記を (1, 1)に変換する
        """
        return coordinate_to_tuple(cell)

    def cells_from_range(self, cell:str) -> Iterator[str]:
        """
        セル範囲を各セルのジェネレーターに変換する
        """
        return chain.from_iterable(cols_from_range(cell))

    def defined_name(self, sheet_name: str | None =None) -> dict:
        """
        定義された名前を取得する
        """
        if sheet_name is None:
            return self.__wb.defined_names
        else:
            return self.__wb[sheet_name].defined_names

    def tables(self) -> dict:
        """
        Excelブックのテーブルを取得する
        """
        tables: dict = {}

        for ws in self.__wb.worksheets:
            for table, ref in ws.tables.items():
                tables |= {
                    table: {
                        "sheet_name": ws.title,
                        "ref": ref
                    }
                }

        return tables
